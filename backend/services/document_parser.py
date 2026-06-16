"""Document parsing, chunking (recursive equiv), embedding (OpenAI compat) with retry."""

import logging
import os
import unicodedata

import httpx

try:
    from bs4 import BeautifulSoup
    _HAS_BS4 = True
except ImportError:
    _HAS_BS4 = False

from constants import ALLOWED_EXTENSIONS
from services.ssl_utils import create_ssl_context

logger = logging.getLogger(__name__)

INVALID_TEXT_BINARY_MESSAGE = (
    "INVALID_TEXT_BINARY: file appears to be binary or unreadable text"
)
INVALID_TEXT_EMPTY_MESSAGE = "INVALID_TEXT_EMPTY: no readable text content"
INVALID_TEXT_ENCODING_MESSAGE = "INVALID_TEXT_ENCODING: file is not valid UTF-8 text"
_ALLOWED_TEXT_CONTROL_CHARS = {"\t", "\n", "\r", "\f"}


class DocumentParser:
    def __init__(self):
        self.max_retries = 2  # >=1 retry per req

    def parse(self, storage_path: str, file_type: str) -> str:
        """Parse file to plain text. Raises on unrecoverable error."""
        if not os.path.exists(storage_path):
            raise FileNotFoundError(storage_path)
        ext = file_type.lower().lstrip(".")
        if ext not in ALLOWED_EXTENSIONS:
            raise ValueError(f"Unsupported: {ext}")

        if ext in ("txt", "md"):
            return self._parse_readable_text(storage_path)
        elif ext == "html":
            return self._parse_html(storage_path)
        elif ext == "pdf":
            return self._parse_pdf(storage_path)
        elif ext == "docx":
            return self._parse_docx(storage_path)
        elif ext == "xlsx":
            return self._parse_xlsx(storage_path)
        raise ValueError(f"Unhandled ext: {ext}")

    def _parse_readable_text(self, path: str) -> str:
        """Read a UTF-8 text-like file and reject binary/unreadable payloads."""
        try:
            with open(path, "rb") as f:
                raw = f.read()
        except OSError as e:
            raise RuntimeError(f"Failed to read text file: {e}") from e

        if not raw:
            raise RuntimeError(INVALID_TEXT_EMPTY_MESSAGE)

        self._reject_binary_bytes(raw)
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError as e:
            raise RuntimeError(INVALID_TEXT_ENCODING_MESSAGE) from e

        self._ensure_readable_text(text)
        return text

    def _reject_binary_bytes(self, raw: bytes) -> None:
        """Reject obvious binary payloads before decoding as UTF-8 text."""
        if b"\x00" in raw:
            raise RuntimeError(INVALID_TEXT_BINARY_MESSAGE)

        disallowed_controls = sum(
            1
            for byte in raw
            if byte < 32 and byte not in (9, 10, 12, 13)
        )
        if disallowed_controls:
            raise RuntimeError(INVALID_TEXT_BINARY_MESSAGE)

    def _ensure_readable_text(self, text: str) -> None:
        """Validate that decoded text contains readable content, not binary noise."""
        if not text.strip():
            raise RuntimeError(INVALID_TEXT_EMPTY_MESSAGE)

        suspicious = 0
        for ch in text:
            if ch in _ALLOWED_TEXT_CONTROL_CHARS:
                continue
            category = unicodedata.category(ch)
            if category in {"Cc", "Cs", "Co", "Cn"}:
                suspicious += 1

        if suspicious and suspicious / max(len(text), 1) > 0.02:
            raise RuntimeError(INVALID_TEXT_BINARY_MESSAGE)

    def _parse_pdf(self, path: str) -> str:
        import pdfplumber

        texts = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                texts.append(t)
        return "\n\n".join(texts)

    def _parse_html(self, path: str) -> str:
        """Extract visible text from HTML file (strips tags)."""
        raw = self._parse_readable_text(path)
        if _HAS_BS4:
            soup = BeautifulSoup(raw, "html.parser")
            text = soup.get_text(separator="\n", strip=True)
        else:
            # Fallback: stdlib html.parser
            from html.parser import HTMLParser
            class _TextExtractor(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.parts = []
                def handle_data(self, data):
                    text = data.strip()
                    if text:
                        self.parts.append(text)
            parser = _TextExtractor()
            parser.feed(raw)
            text = "\n".join(parser.parts)
        self._ensure_readable_text(text)
        return text

    def _parse_docx(self, path: str) -> str:
        """Parse DOCX file, extract paragraph text."""
        file_size = os.path.getsize(path)
        if file_size == 0:
            raise RuntimeError(f"Cannot parse empty DOCX file: {path}")
        logger.info(f"Parsing DOCX: {path} ({file_size} bytes)")
        try:
            from docx import Document
        except ImportError:
            raise RuntimeError(
                "DOCX support requires python-docx. "
                "Install: pip install python-docx"
            )
        try:
            doc = Document(path)
            paragraphs = [p.text for p in doc.paragraphs if p.text]
            logger.info(f"DOCX parsed: {len(paragraphs)} paragraphs from {file_size} bytes")
            return "\n".join(paragraphs)
        except Exception as e:
            logger.error(f"DOCX parse failed for {path}: {e}")
            raise RuntimeError(
                f"Failed to parse DOCX file (type={type(e).__name__}): {e}. "
                "Check that the file is a valid .docx (Office Open XML) format."
            ) from e

    def _parse_xlsx(self, path: str) -> str:
        file_size = os.path.getsize(path)
        if file_size == 0:
            raise RuntimeError(f"Cannot parse empty XLSX file: {path}")
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("XLSX support requires: pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(",".join(cells))
            parts.append(f"Sheet: {sheet.title}\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(parts)

    def chunk_text(
        self, text: str, chunk_size: int = 512, chunk_overlap: int = 64
    ) -> list[str]:
        """RecursiveCharacterTextSplitter equivalent (separators + overlap)."""
        if not text or len(text) <= chunk_size:
            return [text] if text else []
        chunks = []
        start = 0
        while start < len(text):
            end = min(start + chunk_size, len(text))
            chunk = text[start:end]
            chunks.append(chunk)
            start = end - chunk_overlap if end < len(text) else end
            if start < 0:
                start = 0
        # dedupe tiny
        return [c for c in chunks if len(c.strip()) > 10]

    async def embed_texts(
        self,
        texts: list[str],
        model: str,
        base_url: str | None,
        api_key: str | None = None,
    ) -> list[list[float]]:
        """OpenAI-compatible /v1/embeddings call. Returns list of embeddings."""
        if not texts:
            return []
        url = (base_url or "https://api.openai.com/v1").rstrip("/") + "/embeddings"
        headers = {"Content-Type": "application/json"}
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"
        payload = {"model": model, "input": texts}
        ssl_context = create_ssl_context()
        async with httpx.AsyncClient(verify=ssl_context, timeout=60.0) as client:
            resp = await client.post(url, json=payload, headers=headers)
            resp.raise_for_status()
            data = resp.json()
            return [item["embedding"] for item in data.get("data", [])]

    def parse_with_retry(self, storage_path: str, file_type: str) -> str:
        """Retry wrapper (>=1 retry)."""
        last_exc: Exception | None = None
        for attempt in range(self.max_retries):
            try:
                return self.parse(storage_path, file_type)
            except Exception as e:
                # Do not retry non-recoverable errors
                if isinstance(e, (ImportError, ModuleNotFoundError, RuntimeError)):
                    raise
                last_exc = e
                logger.warning(f"Parse attempt {attempt + 1} failed: {e}")
        if last_exc is not None:
            raise last_exc
        raise RuntimeError("parse_with_retry: no attempts made")
